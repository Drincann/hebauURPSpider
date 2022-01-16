import requests_html
import cv2
import threading
import os
from pyquery import PyQuery as jQuery
import openpyxl
import json


class Main:
    # public
    def __init__(this):
        this.session = requests_html.HTMLSession()

    def login(this, user, pwd):
        codeImg = this.__getCodeImg()
        this.__saveTo('./temp.jpg', codeImg)
        data = {"zjh": str(user), "mm": str(pwd), "v_yzm": str(input("验证码："))}
        os.remove('./temp.jpg')
        loginRes = this.session.post(
            'http://urp.hebau.edu.cn/loginAction.do', data=data)
        return True if loginRes.status_code == 200 else False

    def getInfoStructure(this):
        return {
            "name": '',
            "stuid": '',
            "sex": '',
            "id": '',

            "nation": '',
            "nativePlace": '',
            "politicalOutlook": '',
            "birthday": '',

            "clas": '',
            "entrance": '',
            "graduation": '',

            "major": '',

            "Department": '',
            "courses": [],
        }

    def loadInfoById(this, stuid):
        htmlArr = this.__getInfoHtml(stuid)
        stuInfo = {}
        for html in htmlArr:
            jq = jQuery(html)
            # info
            tr = jq('table#report1 tr')
            if jq(jq(tr[1]).children('td')[1]).html() is None:
                return None
            if 'name' not in stuInfo:
                stuInfo = {
                    "name": jq(jq(tr[1]).children('td')[1]).html(),
                    "stuid": jq(jq(tr[1]).children('td')[3]).html(),
                    "sex": jq(jq(tr[1]).children('td')[5]).html(),
                    "id": jq(jq(tr[1]).children('td')[7]).html(),

                    "nation": jq(jq(tr[2]).children('td')[1]).html(),
                    "nativePlace": jq(jq(tr[2]).children('td')[3]).html(),
                    "politicalOutlook": jq(jq(tr[2]).children('td')[5]).html(),
                    "birthday": jq(jq(tr[2]).children('td')[7]).html(),

                    "clas": jq(jq(tr[3]).children('td')[1]).html(),
                    "entrance": jq(jq(tr[3]).children('td')[3]).html(),
                    "graduation": jq(jq(tr[3]).children('td')[5]).html(),

                    "major": jq(jq(tr[4]).children('td')[1]).html(),

                    "Department": jq(jq(tr[5]).children('td')[1]).html(),
                    "courses": [],
                }

            # course
            begin = 7
            end = tr.length - 5
            for i in range(begin, end):
                stuInfo['courses'].append({
                    "name": jq(jq(tr[i]).children('td')[0]).html(),
                    "credit": jq(jq(tr[i]).children('td')[1]).html(),
                    "score": jq(jq(tr[i]).children('td')[2]).html(),
                    "method": jq(jq(tr[i]).children('td')[3]).html(),
                    "attr": jq(jq(tr[i]).children('td')[4]).html(),
                    "time": jq(jq(tr[i]).children('td')[5]).html()
                })
                try:
                    if jq(jq(tr[i]).children('td')[6]).html() is not None:
                        stuInfo['courses'].append({
                            "name": jq(jq(tr[i]).children('td')[6]).html(),
                            "credit": jq(jq(tr[i]).children('td')[7]).html(),
                            "score": jq(jq(tr[i]).children('td')[8]).html(),
                            "method": jq(jq(tr[i]).children('td')[9]).html(),
                            "attr": jq(jq(tr[i]).children('td')[10]).html(),
                            "time": jq(jq(tr[i]).children('td')[11]).html()
                        })
                except Exception as e:
                    pass
        # end for
        return stuInfo

    def getInfoList(this, idRange, saveAsJsonOpts={'save': False, 'dir': ''}):
        infoList = []
        for id in idRange:
            isError = True
            while isError:
                isError = False
                try:
                    info = this.loadInfoById(str(id))
                except Exception as e:
                    isError = True

            if info is not None:
                infoList.append(info)
                print(info['stuid'])
                if saveAsJsonOpts['save'] is True:
                    saveAsJson(infoList, saveAsJsonOpts['dir'])
        return infoList

    # private
    def __getInfoHtml(this, stuid, page=1, toOtherPageData=None):
        if toOtherPageData is None:
            data = {"LS_XH": str(
                stuid), "resultPage": "http://urp.hebau.edu.cn:80/reportFiles/cj/cj_zwcjd.jsp?"}
            res = this.session.post(
                url="http://urp.hebau.edu.cn/setReportParams", data=data)
        else:
            data = toOtherPageData
            res = this.session.post(
                url="http://urp.hebau.edu.cn:80/reportFiles/cj/cj_zwcjd.jsp?", data=data)

        jq = jQuery(res.text)
        buttons = jq('a')
        nextButton = None
        for button in buttons:
            if jq(button).text() == '下一页':
                nextButton = jq(button)
                break
        if nextButton is None:
            return [res.text]
        else:
            nextPageForm = None
            for form in jq('form'):
                if jq(form).attr('name') == 'report1_turnPageForm':
                    nextPageForm = jq(form)
                    break
            if nextPageForm is None:
                return [res.text]
            data = {
                'reportParamsId': nextPageForm.children('input[name="reportParamsId"]').val(),
                'srcId': nextPageForm.children('input[name="srcId"]').val(),
                'report1_currPage': page+1,
                'report1_sessionId': nextPageForm.children('input[name="report1_sessionId"]').val(),
                'report1_cachedId': nextPageForm.children('input[name="report1_cachedId"]').val(),
            }
            returnArr = [res.text]
            returnArr.extend(this.__getInfoHtml(
                stuid, page=page+1, toOtherPageData=data))
            return returnArr

    def __getCodeImg(this):
        return this.session.get('http://urp.hebau.edu.cn/validateCodeAction.do?random=0.08322962004793921').content

    def __saveTo(this, dir, data):
        with open(dir, mode='wb') as f:
            f.write(data)


def saveAsJson(dic, outFile):
    jsStr = json.dumps(dic, ensure_ascii=False)
    with open(outFile, 'wt', encoding='utf-8') as f:
        f.write(jsStr)


def get1_5InMem(obj):
    idRange = list(range(2019984040101, 2019984040131)) \
        + list(range(2019984040201, 2019984040231)) \
        + list(range(2019984040301, 2019984040331)) \
        + list(range(2019984040401, 2019984040431)) \
        + list(range(2019984040501, 2019984040531))
    return obj.getInfoList(idRange)


def getRangeOfSG():
    return list(range(2019984020101, 2019984020124)) \
        + list(range(2019984020201, 2019984020224)) \
        + list(range(2019984020301, 2019984020321)) \
        + list(range(2019984020401, 2019984020424)) \
        + list(range(2019984020501, 2019984020525))


def get1_5RangeOfSE():
    return list(range(2019984040101, 2019984040131)) \
        + list(range(2019984040201, 2019984040231)) \
        + list(range(2019984040301, 2019984040331)) \
        + list(range(2019984040401, 2019984040431)) \
        + list(range(2019984040501, 2019984040531))


def get1_5RangeOfCS():
    return \
        list(range(2019984130101, 2019984130132)) \
        + list(range(2019984130201, 2019984130231)) \
        + list(range(2019984130301, 2019984130331)) \
        + list(range(2019984130401, 2019984130430)) \
        + list(range(2019984130501, 2019984130530))\
        + [2019984010119, 2019984010206, 2019984110409, 2019984110429, 2017984130409,
            2019984020303, 2019984020305, 2019984010211, 2019984010308, 2019984010527]


def getRangeOfKJ():
    return list(range(2019994090101, 2019994090131)) \
        + list(range(2019994090201, 2019994090231)) \
        + list(range(2019994090301, 2019994090330)) \
        + list(range(2019994090401, 2019994090430)) \
        + list(range(2019994090501, 2019994090529)) \
        + list(range(2019994090601, 2019994090630)) \
        + [2019994020225, 2019994020224, 2019994020229, 2019994020210, 2019994130129, 2019994130126, 2019994130130, 2019994130128, 2019994130215, 2019994130210, 2019994130228, 2019984080418, 2019984080410, 2019984020127, 2019984020121, 2019994020220, 2019984020128, 2019984020310,
            2019984020319, 2019984020313, 2019984020323, 2019984020409, 2019984020410, 2019994070229, 2019994070211, 2019994070333, 2019994070313, 2019994070431, 2019994070428, 2019984020109, 2019994070527, 2019994070130, 2019994070629, 2019994130201, 2019994130105, 2019994070203]
# 去除正常考试及补考的信息


def duplicateRemoval(infoList):
    for info in infoList:
        refresher = []
        makeUp = []

        # 若复修了，则删掉重修、正常以及补考的的
        for course in info['courses']:
            if course['method'] == '复修':
                refresher.append(course['name'])
        if len(refresher) != 0:
            for course in info['courses']:
                if course['method'] == '重修' and course['name'] in refresher:
                    info['courses'].remove(course)
                if course['method'] == '正常' and course['name'] in refresher:
                    info['courses'].remove(course)
                if course['method'] == '补考' and course['name'] in refresher:
                    info['courses'].remove(course)

        refresher = []
        # 仅保留最新一次考试，所以若重修了则需要把补考和正常的删掉，
        for course in info['courses']:
            if course['method'] == '重修':
                refresher.append(course['name'])
        if len(refresher) != 0:
            for course in info['courses']:
                if course['method'] == '正常' and course['name'] in refresher:
                    info['courses'].remove(course)
                if course['method'] == '补考' and course['name'] in refresher:
                    info['courses'].remove(course)

        # 若未重修，但补考了，则需要删掉正常的
        for course in info['courses']:
            if course['method'] == '补考':
                makeUp.append(course['name'])
        if len(makeUp) != 0:
            for course in info['courses']:
                if course['method'] == '正常' and course['name'] in makeUp:
                    info['courses'].remove(course)
    return infoList


def calGPA(info):
    scoreSum = 0.0
    creditSum = 0.0
    for course in info['courses']:
        try:
            score = float(course['credit']) * float(course['score'])
            credit = float(course['credit'])
            scoreSum += score
            creditSum += credit
        except Exception as e:
            print(e)
    return scoreSum / creditSum


def calGPAArray(infoList):
    GPAArr = []
    for info in infoList:
        try:
            GPAArr.append({
                "name": info['name'],
                "stuid": info['stuid'],
                "GPA": calGPA(info),
            })
        except Exception as e:
            GPAArr.append({
                "name": info['name'],
                "stuid": info['stuid'],
                "GPA": 0,
            })
            print(f'{info["stuid"]} {info["name"]} {e}')

    return GPAArr


def sortedGPAList(infoList):
    GPAArr = calGPAArray(infoList)
    return sorted(GPAArr, key=lambda x: x['GPA'], reverse=True)

# 爬出成绩并 save as json


def getAndSaveAsJson(uname, pwd, idRange, outSrc, outCleaned):
    obj = Main()
    if not (obj.login(uname, pwd)):
        exit()
    infoList = obj.getInfoList(idRange, saveAsJsonOpts={
                               'save': True, 'dir': outSrc})
    infoList = duplicateRemoval(infoList)
    saveAsJson(infoList, outCleaned)
    return infoList

# 姓名学号学分成绩和成绩
# scoreTplStuid 是以谁的科目做显示模板的那个人的学号


def saveAsXlsx(infoList, scoreTplStuid, outFile):
    # 为了在后面方便地拿到科目成绩，这里需要散列一下，弄一个 coursesDic
    coursesDic = {}
    for info in infoList:
        courseDic = {}
        for course in info['courses']:
            courseDic[course['name']] = course
        coursesDic[info['stuid']] = courseDic
    tplCourses = coursesDic[scoreTplStuid]
    for info in infoList:
        tempCourseList = info['courses']
        info['courses'] = []
        for course in tempCourseList:
            if course['name'] in tplCourses:
                info['courses'].append(course)

    # 排序过的 infoList
    sortedList = sortedGPAList(infoList)

    # init excel
    wb = openpyxl.Workbook()
    ws = wb.active

    row = 1
    col = 1
    for key in sortedList[0]:
        ws.cell(row=row, column=col).value = key
        col += 1
    courses = []
    for course in tplCourses:
        courses.append(tplCourses[course]['name'])
        ws.cell(row=row, column=col).value = tplCourses[course]['name']
        col += 1
        pass
    row += 1
    for info in sortedList:
        col = 1
        for key in info.keys():
            ws.cell(row=row, column=col).value = info[key]
            col += 1
        for key in courses:
            if key not in coursesDic[info['stuid']]:
                ws.cell(row=row, column=col).value = ''
            else:
                ws.cell(
                    row=row, column=col).value = coursesDic[info['stuid']][key]['score']
            col += 1
        row += 1
        print(info)
    wb.save(outFile)


if __name__ == '__main__':
    # example
    infoList = getAndSaveAsJson('username', 'password',
                                get1_5RangeOfCS(), './cssrc.json', './cscleaned.json')
    saveAsXlsx(infoList, '2019984130206', '/Users/coder/Desktop/outCs.xlsx')
